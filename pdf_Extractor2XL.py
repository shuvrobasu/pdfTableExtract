import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
import pdfplumber
from openpyxl import Workbook

HANDLE_RADIUS = 12

class PDFTableSelectorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PDF Table Selector and Exporter")
        self.root.protocol("WM_DELETE_WINDOW", self.on_exit)
        self.pdf_folder = ""
        self.all_pdf_files = []  # all found in folder
        self.pdf_files = []      # filtered
        self.selected_pdf = None
        self.current_page = 0
        self.total_pages = 0
        self.tk_img = None
        self.tables_on_page = []
        self.table_bboxes = []
        self.selected_tables = set()
        self.all_tables = {}
        self.preview_img = None
        self.pdf_page_width = 1
        self.pdf_page_height = 1
        self.handle_coords = []
        self.is_loading = False

        self.DPI = 200
        self.dpi_scale = self.DPI / 72
        self.border = 60

        self.init_gui()

    def init_gui(self):
        frm = ttk.Frame(self.root, padding=10)
        frm.grid(sticky="nsew")
        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)
        frm.rowconfigure(1, weight=1)

        # ---- Filter + Folder select ----
        filter_frame = ttk.Frame(frm)
        filter_frame.grid(row=0, column=0, sticky="ew")
        filter_frame.columnconfigure(0, weight=1)

        self.filter_var = tk.StringVar()
        self.filter_entry = ttk.Entry(filter_frame, textvariable=self.filter_var)
        self.filter_entry.grid(row=0, column=0, sticky="ew", padx=(0,4))
        self.filter_entry.bind("<KeyRelease>", self.on_filter_change)
        self.filter_entry.bind("<Return>", self.on_filter_change)

        btn_select_folder = ttk.Button(filter_frame, text="Select PDF Folder", command=self.select_folder)
        btn_select_folder.grid(row=0, column=1, sticky="e", padx=(0,2))

        # ---- Listbox ----
        self.listbox = tk.Listbox(frm, width=40)
        self.listbox.grid(row=1, column=0, sticky="ns", pady=2)
        self.listbox.bind('<<ListboxSelect>>', self.on_pdf_select)

        self.scrollbar = ttk.Scrollbar(frm, orient="vertical", command=self.listbox.yview)
        self.scrollbar.grid(row=1, column=0, sticky="nse", padx=(0,2))
        self.listbox.config(yscrollcommand=self.scrollbar.set)

        # ======= Preview Panel with Scrollbars =======
        self.img_panel = ttk.Frame(frm)
        self.img_panel.grid(row=1, column=1, sticky="nsew", padx=(10,0))
        self.img_panel.columnconfigure(0, weight=1)
        self.img_panel.rowconfigure(0, weight=1)

        self.canvas_frame = ttk.Frame(self.img_panel)
        self.canvas_frame.grid(row=0, column=0, sticky="nsew")
        self.canvas_frame.columnconfigure(0, weight=1)
        self.canvas_frame.rowconfigure(0, weight=1)

        # Scrollbars
        self.xscroll = tk.Scrollbar(self.canvas_frame, orient="horizontal")
        self.xscroll.grid(row=1, column=0, sticky="ew")
        self.yscroll = tk.Scrollbar(self.canvas_frame, orient="vertical")
        self.yscroll.grid(row=0, column=1, sticky="ns")

        self.canvas = tk.Canvas(
            self.canvas_frame,
            bg="white",
            highlightthickness=0,
            xscrollcommand=self.xscroll.set,
            yscrollcommand=self.yscroll.set,
            width=1200, height=1400
        )
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.canvas.bind("<Button-1>", self.on_canvas_click)
        self.xscroll.config(command=self.canvas.xview)
        self.yscroll.config(command=self.canvas.yview)

        # Page controls for multi-page preview
        self.page_control_frame = ttk.Frame(self.img_panel)
        self.page_control_frame.grid(row=1, column=0, sticky="ew", pady=3)
        self.btn_prev = ttk.Button(self.page_control_frame, text="<< Prev", command=self.prev_page, state="disabled")
        self.btn_next = ttk.Button(self.page_control_frame, text="Next >>", command=self.next_page, state="disabled")
        self.page_label = ttk.Label(self.page_control_frame, text="Page 1/1")
        self.btn_prev.grid(row=0, column=0)
        self.page_label.grid(row=0, column=1, padx=8)
        self.btn_next.grid(row=0, column=2)

        # Table controls
        self.table_control_frame = ttk.Frame(self.img_panel)
        self.table_control_frame.grid(row=2, column=0, sticky="ew", pady=5)
        self.btn_select_all = ttk.Button(self.table_control_frame, text="Select All Tables", command=self.select_all_tables, state="disabled")
        self.btn_export_selected = ttk.Button(self.table_control_frame, text="Export Selected Tables", command=self.export_selected_tables, state="disabled")
        self.btn_export_all = ttk.Button(self.table_control_frame, text="Export All Tables", command=self.export_all_tables, state="disabled")
        self.btn_select_all.grid(row=0, column=0, padx=2)
        self.btn_export_selected.grid(row=0, column=1, padx=2)
        self.btn_export_all.grid(row=0, column=2, padx=2)

        # Exit button at bottom right
        self.exit_bar = ttk.Frame(self.root)
        self.exit_bar.grid(row=2, column=0, sticky="ew", pady=(5,2), padx=(0,10))
        self.exit_bar.columnconfigure(0, weight=1)
        self.btn_exit = ttk.Button(self.exit_bar, text="Exit", command=self.on_exit)
        self.btn_exit.grid(row=0, column=1, sticky="e", padx=5)

        self.status_var = tk.StringVar()
        self.status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor="w", padding=3)
        self.status_bar.grid(row=3, column=0, sticky="ew")

    def on_exit(self):
        self.root.quit()
        self.root.destroy()

    def set_status(self, text):
        self.status_var.set(text)
        self.root.update_idletasks()

    def select_folder(self):
        folder = filedialog.askdirectory(title="Select Folder with PDF files")
        if folder:
            self.pdf_folder = folder
            self.load_pdfs()

    def load_pdfs(self):
        self.all_pdf_files = [f for f in os.listdir(self.pdf_folder) if f.lower().endswith(".pdf")]
        self.filter_var.set("")  # reset filter on reload
        self.apply_filter()
        self.selected_pdf = None
        self.current_page = 0
        self.total_pages = 0
        self.tables_on_page = []
        self.table_bboxes = []
        self.selected_tables.clear()
        self.all_tables = {}
        self.btn_export_selected.config(state="disabled")
        self.btn_export_all.config(state="disabled")
        self.btn_select_all.config(state="disabled")
        self.canvas.delete("all")
        self.set_status(f"Found {len(self.all_pdf_files)} PDF files.")

    def apply_filter(self):
        """Filters self.all_pdf_files to self.pdf_files and refreshes listbox."""
        value = self.filter_var.get().strip().lower()
        if value:
            self.pdf_files = [f for f in self.all_pdf_files if value in f.lower()]
        else:
            self.pdf_files = list(self.all_pdf_files)
        self.listbox.delete(0, tk.END)
        for pdf in self.pdf_files:
            self.listbox.insert(tk.END, pdf)
        if not self.pdf_files:
            self.selected_pdf = None

    def on_filter_change(self, event=None):
        self.apply_filter()
        # Optionally select first file automatically
        if self.pdf_files:
            self.listbox.selection_set(0)
            self.listbox.activate(0)
            self.listbox.see(0)
        else:
            self.listbox.selection_clear(0, tk.END)

    def on_pdf_select(self, event):
        sel = self.listbox.curselection()
        if sel:
            idx = sel[0]
            pdf_name = self.pdf_files[idx]
            self.selected_pdf = pdf_name
            self.current_page = 0
            self.selected_tables.clear()
            self.all_tables.clear()
            self.is_loading = True
            self.set_status("Loading tables, please wait...")
            self.disable_controls()
            threading.Thread(target=self.preload_all_tables, args=(os.path.join(self.pdf_folder, pdf_name),), daemon=True).start()
        else:
            self.selected_pdf = None
            self.canvas.delete("all")
            self.tables_on_page = []
            self.table_bboxes = []
            self.selected_tables.clear()
            self.all_tables.clear()
            self.btn_export_selected.config(state="disabled")
            self.btn_export_all.config(state="disabled")
            self.btn_select_all.config(state="disabled")
            self.update_page_controls()

    def disable_controls(self):
        self.btn_export_selected.config(state="disabled")
        self.btn_export_all.config(state="disabled")
        self.btn_select_all.config(state="disabled")
        self.btn_prev.config(state="disabled")
        self.btn_next.config(state="disabled")
        self.listbox.config(state="disabled")
        self.btn_exit.config(state="disabled")
        self.filter_entry.config(state="disabled")

    def enable_controls(self):
        self.btn_export_all.config(state="normal" if self.all_tables else "disabled")
        self.btn_select_all.config(state="normal" if self.tables_on_page else "disabled")
        self.btn_export_selected.config(state="normal" if self.selected_tables else "disabled")
        self.btn_prev.config(state="normal" if self.current_page > 0 else "disabled")
        self.btn_next.config(state="normal" if self.current_page < self.total_pages-1 else "disabled")
        self.listbox.config(state="normal")
        self.btn_exit.config(state="normal")
        self.filter_entry.config(state="normal")

    def preload_all_tables(self, pdf_path):
        self.all_tables = {}
        try:
            with pdfplumber.open(pdf_path) as pdf:
                self.total_pages = len(pdf.pages)
                for page_num, page in enumerate(pdf.pages):
                    tables, bboxes = [], []
                    for t in page.find_tables():
                        tables.append(t.extract())
                        bboxes.append(t.bbox)
                    self.all_tables[page_num] = list(zip(tables, bboxes))
        except Exception as e:
            self.root.after(0, lambda: self.set_status(f"Failed to read tables: {e}"))
        self.root.after(0, self.after_preload_all_tables)

    def after_preload_all_tables(self):
        self.display_pdf_with_tables(os.path.join(self.pdf_folder, self.selected_pdf), page_number=0)
        self.update_page_controls()
        self.is_loading = False
        self.enable_controls()

    def display_pdf_with_tables(self, pdf_path, page_number=0):
        self.tables_on_page = []
        self.table_bboxes = []
        self.handle_coords = []
        self.canvas.delete("all")
        try:
            with pdfplumber.open(pdf_path) as pdf:
                page = pdf.pages[page_number]
                self.pdf_page_width, self.pdf_page_height = page.width, page.height
                img = page.to_image(resolution=self.DPI).original
                bg_img = Image.new("RGB", (img.width + 2*self.border, img.height + 2*self.border), "white")
                bg_img.paste(img, (self.border, self.border))
                self.preview_img = bg_img
                self.tk_img = ImageTk.PhotoImage(bg_img)
                self.canvas.config(scrollregion=(0,0, bg_img.width, bg_img.height))
                self.canvas.config(width=min(bg_img.width, 1200), height=min(bg_img.height, 1400))
                self.canvas.create_image(0, 0, anchor=tk.NW, image=self.tk_img)
                if page_number in self.all_tables:
                    self.tables_on_page = [t for t, b in self.all_tables[page_number]]
                    self.table_bboxes = [b for t, b in self.all_tables[page_number]]
                else:
                    self.tables_on_page = []
                    self.table_bboxes = []
                self.redraw_table_bboxes()
                self.btn_select_all.config(state="normal" if self.tables_on_page else "disabled")
                self.btn_export_selected.config(state="normal" if self.selected_tables else "disabled")
                self.set_status(f"Previewing: {os.path.basename(pdf_path)}, page {page_number+1}/{self.total_pages}. Click handles or inside rectangles to select.")
        except Exception as e:
            self.set_status(f"Preview failed: {e}")

    def update_page_controls(self):
        if self.total_pages > 1:
            self.btn_prev.config(state="normal" if self.current_page > 0 else "disabled")
            self.btn_next.config(state="normal" if self.current_page < self.total_pages-1 else "disabled")
        else:
            self.btn_prev.config(state="disabled")
            self.btn_next.config(state="disabled")
        self.page_label.config(text=f"Page {self.current_page+1 if self.total_pages else 0}/{self.total_pages if self.total_pages else 1}")

    def prev_page(self):
        if self.selected_pdf and self.current_page > 0:
            self.current_page -= 1
            self.display_pdf_with_tables(os.path.join(self.pdf_folder, self.selected_pdf), self.current_page)
            self.update_page_controls()

    def next_page(self):
        if self.selected_pdf and self.current_page < self.total_pages-1:
            self.current_page += 1
            self.display_pdf_with_tables(os.path.join(self.pdf_folder, self.selected_pdf), self.current_page)
            self.update_page_controls()

    def draw_bbox(self, bbox, outline="red", width=5, fill=None, idx=None):
        x0, y0, x1, y1 = bbox
        x0_img = x0 * self.dpi_scale + self.border
        y0_img = y0 * self.dpi_scale + self.border
        x1_img = x1 * self.dpi_scale + self.border
        y1_img = y1 * self.dpi_scale + self.border
        self.canvas.create_rectangle(x0_img, y0_img, x1_img, y1_img, outline=outline, width=5, fill=fill, tags="table_bbox")
        handle_color = "green" if outline == "blue" else "orange"
        corners = [
            (x0_img, y0_img),
            (x1_img, y0_img),
            (x1_img, y1_img),
            (x0_img, y1_img)
        ]
        for corner_idx, (cx, cy) in enumerate(corners):
            self.canvas.create_oval(cx-HANDLE_RADIUS, cy-HANDLE_RADIUS, cx+HANDLE_RADIUS, cy+HANDLE_RADIUS,
                                   fill=handle_color, outline="black", width=2, tags="handle")
            self.handle_coords.append((idx, corner_idx, cx, cy))

    def get_canvas_bbox_coords(self, bbox):
        x0, y0, x1, y1 = bbox
        x0_img = x0 * self.dpi_scale + self.border
        y0_img = y0 * self.dpi_scale + self.border
        x1_img = x1 * self.dpi_scale + self.border
        y1_img = y1 * self.dpi_scale + self.border
        return (x0_img, y0_img, x1_img, y1_img)

    def redraw_table_bboxes(self):
        self.canvas.delete("table_bbox")
        self.canvas.delete("handle")
        self.handle_coords = []
        for idx, bbox in enumerate(self.table_bboxes):
            key = (self.current_page, idx)
            color = "blue" if key in self.selected_tables else "red"
            self.draw_bbox(bbox, outline=color, width=5, idx=idx)

    def on_canvas_click(self, event):
        if not self.table_bboxes:
            return
        click_x, click_y = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        for idx, corner_idx, cx, cy in self.handle_coords:
            if (click_x - cx) ** 2 + (click_y - cy) ** 2 <= HANDLE_RADIUS ** 2:
                key = (self.current_page, idx)
                if key in self.selected_tables:
                    self.selected_tables.remove(key)
                else:
                    self.selected_tables.add(key)
                self.redraw_table_bboxes()
                self.btn_export_selected.config(state="normal" if self.selected_tables else "disabled")
                if self.selected_tables:
                    self.set_status(f"{len(self.selected_tables)} table(s) selected. Export or select more.")
                else:
                    self.set_status("Click handles or inside rectangles to select.")
                return
        for idx, bbox in enumerate(self.table_bboxes):
            rect = self.get_canvas_bbox_coords(bbox)
            if rect[0] <= click_x <= rect[2] and rect[1] <= click_y <= rect[3]:
                key = (self.current_page, idx)
                if key in self.selected_tables:
                    self.selected_tables.remove(key)
                else:
                    self.selected_tables.add(key)
                self.redraw_table_bboxes()
                self.btn_export_selected.config(state="normal" if self.selected_tables else "disabled")
                if self.selected_tables:
                    self.set_status(f"{len(self.selected_tables)} table(s) selected. Export or select more.")
                else:
                    self.set_status("Click handles or inside rectangles to select.")
                return

    def select_all_tables(self):
        self.selected_tables = set((self.current_page, idx) for idx in range(len(self.table_bboxes)))
        self.redraw_table_bboxes()
        self.btn_export_selected.config(state="normal" if self.selected_tables else "disabled")
        self.set_status(f"All {len(self.selected_tables)} tables on this page selected.")

    def export_selected_tables(self):
        if not self.selected_tables:
            messagebox.showinfo("No Selection", "Select tables first.")
            return
        pdf_path = os.path.join(self.pdf_folder, self.selected_pdf)
        out_path = os.path.splitext(pdf_path)[0] + "_selected_tables.xlsx"
        wb = Workbook()
        wb.remove(wb.active)
        count = 0
        for page_num, idx in sorted(self.selected_tables):
            try:
                table, bbox = self.all_tables[page_num][idx]
            except Exception:
                continue
            if table:
                ws = wb.create_sheet(title=f"Page{page_num+1}_Table{idx+1}")
                for row in table:
                    ws.append(row)
                count += 1
        if count == 0:
            messagebox.showinfo("No Data", "No data to export.")
            return
        wb.save(out_path)
        messagebox.showinfo("Exported", f"{count} table(s) exported to:\n{out_path}")
        self.set_status(f"Exported {count} selected table(s) to {os.path.basename(out_path)}")

    def export_all_tables(self):
        if not self.all_tables:
            messagebox.showinfo("No Tables", "No tables found.")
            return
        pdf_path = os.path.join(self.pdf_folder, self.selected_pdf)
        out_path = os.path.splitext(pdf_path)[0] + "_all_tables.xlsx"
        wb = Workbook()
        wb.remove(wb.active)
        count = 0
        for page_num, tables in self.all_tables.items():
            for idx, (table, bbox) in enumerate(tables):
                if table:
                    ws = wb.create_sheet(title=f"Page{page_num+1}_Table{idx+1}")
                    for row in table:
                        ws.append(row)
                    count += 1
        if count == 0:
            messagebox.showinfo("No Data", "No data to export.")
            return
        wb.save(out_path)
        messagebox.showinfo("Exported", f"{count} table(s) exported to:\n{out_path}")
        self.set_status(f"Exported {count} table(s) to {os.path.basename(out_path)}")

def main():
    root = tk.Tk()
    app = PDFTableSelectorApp(root)
    root.geometry("1500x1100")
    root.mainloop()

if __name__ == '__main__':
    main()
